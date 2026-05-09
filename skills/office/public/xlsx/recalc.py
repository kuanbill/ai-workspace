#!/usr/bin/env python3
"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import json
import sys
import subprocess
import os
import platform
import tempfile
from pathlib import Path
from openpyxl import load_workbook


def recalc(filename, timeout=30, soffice_cmd='soffice'):
    """
    Recalculate formulas in Excel file and report any errors
    
    Uses LibreOffice --convert-to to trigger recalculation, which is cross-platform
    and avoids UNO macro issues on Windows.
    
    Args:
        filename: Path to Excel file
        timeout: Maximum time to wait for recalculation (seconds)
        soffice_cmd: Path to LibreOffice soffice executable
    
    Returns:
        dict with error locations and counts
    """
    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}
    
    abs_path = str(Path(filename).absolute())
    output_dir = Path(tempfile.mkdtemp(prefix="xlsx_recalc_"))
    
    try:
        # Use --convert-to to force LibreOffice to recalculate formulas
        cmd = [
            soffice_cmd, '--headless', '--convert-to', 'xlsx',
            '--outdir', str(output_dir),
            abs_path,
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        
        if result.returncode != 0:
            error_msg = (result.stderr or result.stdout or 'Unknown error').strip()
            return {'error': f'LibreOffice recalculation failed: {error_msg}'}
        
        # Replace the original file with the recalculated version
        converted = output_dir / Path(filename).name
        if converted.exists():
            import shutil
            shutil.copy2(str(converted), abs_path)
        else:
            return {'error': 'Recalculated file not produced'}
    
    except subprocess.TimeoutExpired:
        return {'error': 'LibreOffice recalculation timed out'}
    except Exception as e:
        return {'error': str(e)}
    finally:
        import shutil
        shutil.rmtree(str(output_dir), ignore_errors=True)
    
    # Check for Excel errors in the recalculated file - scan ALL cells
    try:
        wb = load_workbook(filename, data_only=True)
        
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Check ALL rows and columns - no limits
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break
        
        wb.close()
        
        # Build result summary
        result = {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'error_summary': {}
        }
        
        # Add non-empty error categories
        for err_type, locations in error_details.items():
            if locations:
                result['error_summary'][err_type] = {
                    'count': len(locations),
                    'locations': locations[:20]  # Show up to 20 locations
                }
        
        # Add formula count for context - also check ALL cells
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        wb_formulas.close()
        
        result['total_formulas'] = formula_count
        
        return result
        
    except Exception as e:
        return {'error': str(e)}


def main():
    # Parse --soffice-path from args, filter it out for positional args
    soffice_path = None
    positional = []
    i = 1
    while i < len(sys.argv):
        if sys.argv[i] == '--soffice-path' and i + 1 < len(sys.argv):
            soffice_path = sys.argv[i + 1]
            i += 2
        else:
            positional.append(sys.argv[i])
            i += 1

    if len(positional) < 1:
        print("Usage: python recalc.py <excel_file> [timeout_seconds] [--soffice-path PATH]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)
    
    filename = positional[0]
    timeout = int(positional[1]) if len(positional) > 1 else 30
    soffice_cmd = soffice_path or 'soffice'
    
    result = recalc(filename, timeout, soffice_cmd)
    print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()